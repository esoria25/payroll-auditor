#!/usr/bin/env python3
import pandas as pd
import sys
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# Define colors
RED_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
GREEN_FILL = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
BOLD_FONT = Font(bold=True, color="FFFFFF")
BOLD_BLACK = Font(bold=True)

def generate_custom_excel_report(file1_path, file2_path, output_excel_path):
    print(f"\n{'='*80}")
    print(f"GENERATING CUSTOM EXCEL REPORT: {output_excel_path}")
    print(f"File 1: {file1_path}")
    print(f"File 2: {file2_path}")
    print(f"{'='*80}\n")

    try:
        from universal_payroll_auditor import UniversalPayrollAuditor
        auditor = UniversalPayrollAuditor()
        audit_result = auditor.audit(file1_path, file2_path)
    except Exception as e:
        print(f"Error running audit: {e}")
        print("Please ensure universal_payroll_auditor.py is correctly set up.")
        sys.exit(1)

    # Create Excel writer
    with pd.ExcelWriter(output_excel_path, engine='openpyxl') as writer:
        
        # SHEET 1: Summary
        summary_data = {
            'Metric': [
                'Total Rows Compared', 'Matching Rows', 'Different Rows',
                'Match Percentage', 'Total Differences', 'Comparison Date',
                'File 1', 'File 2'
            ],
            'Value': [
                audit_result['summary'].get('total_rows', 0),
                audit_result['summary'].get('matching_rows', 0),
                audit_result['summary'].get('different_rows', 0),
                f"{audit_result['summary'].get('match_percentage', 0.0):.2f}%",
                audit_result['summary'].get('total_differences', 0),
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                file1_path.split('/')[-1],
                file2_path.split('/')[-1]
            ]
        }
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        
        # Format Summary sheet
        ws_summary = writer.sheets['Summary']
        for cell in ws_summary[1]:
            cell.fill = HEADER_FILL
            cell.font = BOLD_FONT
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 50
        
        # SHEET 2: Side-by-Side Comparison (Option 1 format)
        if audit_result.get('differences'):
            # Create side-by-side comparison rows
            comparison_rows = []
            
            for diff_entry in audit_result.get('differences', []):
                employee_name = diff_entry.get('employee', 'N/A')
                
                for field_diff in diff_entry.get('fields', []):
                    field_name = field_diff['field']
                    file1_value = field_diff['file1']
                    file2_value = field_diff['file2']
                    difference = field_diff.get('difference', 'N/A')
                    
                    comparison_rows.append({
                        'Employee': employee_name,
                        'Field': field_name,
                        'File 1 Value': file1_value,
                        'File 2 Value': file2_value,
                        'Difference': difference,
                        'Status': 'DISCREPANCY'
                    })
            
            if comparison_rows:
                comparison_df = pd.DataFrame(comparison_rows)
                comparison_df.to_excel(writer, sheet_name='Side-by-Side Comparison', index=False)
                
                # Format Side-by-Side sheet with RED highlighting
                ws_comparison = writer.sheets['Side-by-Side Comparison']
                
                # Header formatting
                for cell in ws_comparison[1]:
                    cell.fill = HEADER_FILL
                    cell.font = BOLD_FONT
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Highlight discrepancies in RED
                file1_col = comparison_df.columns.get_loc('File 1 Value') + 1
                file2_col = comparison_df.columns.get_loc('File 2 Value') + 1
                status_col = comparison_df.columns.get_loc('Status') + 1
                
                for row_idx in range(2, len(comparison_df) + 2):
                    # Highlight File 1 and File 2 values in RED
                    ws_comparison.cell(row=row_idx, column=file1_col).fill = RED_FILL
                    ws_comparison.cell(row=row_idx, column=file2_col).fill = RED_FILL
                    ws_comparison.cell(row=row_idx, column=file1_col).font = BOLD_BLACK
                    ws_comparison.cell(row=row_idx, column=file2_col).font = BOLD_BLACK
                    
                    # Status column
                    status_cell = ws_comparison.cell(row=row_idx, column=status_col)
                    status_cell.fill = RED_FILL
                    status_cell.font = BOLD_BLACK
                
                # Auto-adjust column widths
                for column in ws_comparison.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws_comparison.column_dimensions[column_letter].width = adjusted_width
        
        # SHEET 3: Matching Rows
        if audit_result.get('matching_rows'):
            matching_df = pd.DataFrame(audit_result['matching_rows'])
            matching_df['Status'] = 'MATCH'
            matching_df.to_excel(writer, sheet_name='Matching Rows', index=False)
            
            ws_matching = writer.sheets['Matching Rows']
            for cell in ws_matching[1]:
                cell.fill = HEADER_FILL
                cell.font = BOLD_FONT
            
            # Highlight matching rows in GREEN
            for row_idx in range(2, len(matching_df) + 2):
                for col_idx in range(1, len(matching_df.columns) + 1):
                    ws_matching.cell(row=row_idx, column=col_idx).fill = GREEN_FILL
            
            for column in ws_matching.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_matching.column_dimensions[column_letter].width = adjusted_width

    print(f"✅ Custom Excel report created: {output_excel_path}")
    print(f"   📊 Sheets: Summary | Side-by-Side Comparison (RED highlights) | Matching Rows (GREEN)")
    print(f"{'='*80}\n")
    return output_excel_path

if __name__ == '__main__':
    if len(sys.argv) < 3:
        print("Usage: python3 create_excel_report.py file1_path file2_path [output_excel_path]")
        sys.exit(1)
    
    file1 = sys.argv[1]
    file2 = sys.argv[2]
    output = sys.argv[3] if len(sys.argv) > 3 else 'payroll_audit_report_custom.xlsx'
    
    generate_custom_excel_report(file1, file2, output)

